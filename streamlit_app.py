# ======================================================================
# FILE 1: streamlit_tabel_utama.py  (Rekonsiliasi Payment Report)
# ======================================================================
import io
import zipfile
from datetime import date
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict, OrderedDict
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook  # streaming Excel

# Optional akselerasi CSV
_HAS_PYARROW = False
try:
    import pyarrow  # noqa: F401
    _HAS_PYARROW = True
except Exception:
    _HAS_PYARROW = False


# =========================== Konfigurasi ===========================
COL_H = "TIPE PEMBAYARAN"                      # H
COL_B = "TANGGAL PEMBAYARAN"                   # B
COL_AA = "REF NO"                              # AA
COL_K = "TOTAL TARIF TANPA BIAYA ADMIN (Rp.)"  # K
COL_X = "SOF ID"                               # X
COL_ASAL = "ASAL"                              # ASAL/Pelabuhan
REQUIRED_COLS = [COL_H, COL_B, COL_AA, COL_K, COL_X, COL_ASAL]

CAT_COLS = [
    "Cash", "Prepaid BRI", "Prepaid BNI", "Prepaid Mandiri", "Prepaid BCA",
    "SKPT", "IFCS", "Reedem", "ESPAY", "Finnet",
]
NON_COMPONENTS = ["Cash", "Prepaid BRI", "Prepaid BNI", "Prepaid Mandiri", "Prepaid BCA", "SKPT", "IFCS", "Reedem"]

CSV_CHUNK_ROWS = 300_000
XLSX_BATCH_ROWS = 80_000
VALID_EXTS = (".xlsx", ".xls", ".csv")


# =========================== Utilitas ===========================
def _style_table(df_display: pd.DataFrame, highlight: bool = False) -> "pd.io.formats.style.Styler":
    numeric_cols = df_display.select_dtypes(include="number").columns.tolist()
    styler = df_display.style.format("{:,.0f}", subset=numeric_cols)
    if highlight and "Selisih" in df_display.columns:
        styler = styler.apply(
            lambda s: [
                "background-color:#fdecea; color:#b71c1c; font-weight:600;" if (pd.notna(v) and float(v) != 0) else ""
                for v in s
            ],
            subset=["Selisih"],
        )
    return styler

def _add_subtotal_row(df_display: pd.DataFrame, label: str = "Subtotal", date_col: str = "Tanggal") -> pd.DataFrame:
    # why: subtotal di UI
    numeric_cols = df_display.select_dtypes(include="number").columns.tolist()
    totals = df_display[numeric_cols].sum()
    subtotal = {c: (totals[c] if c in totals else None) for c in df_display.columns}
    subtotal[date_col] = label
    return pd.concat([df_display, pd.DataFrame([subtotal])], ignore_index=True)

def _to_excel_bytes(df: pd.DataFrame, sheet_name: str) -> Tuple[Optional[bytes], Optional[str], Optional[str]]:
    # why: fallback 2 engine untuk portabilitas
    for engine in ("xlsxwriter", "openpyxl"):
        try:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine=engine) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            return buf.getvalue(), engine, None
        except ImportError:
            continue
        except Exception as e:
            return None, None, f"Gagal menulis Excel dengan {engine}: {e}"
    return None, None, "Tidak ada engine Excel (xlsxwriter/openpyxl)."

def _empty_agg_key2():
    return defaultdict(lambda: defaultdict(float))  # key: (Tanggal, Key2) -> {col -> sum}

def _merge_aggs(a: Dict, b: Dict) -> Dict:
    for k, bucket in b.items():
        for col, val in bucket.items():
            a[k][col] += val
    return a

def _update_agg_series_key2(agg, ser: pd.Series, colname: str) -> None:
    if ser.empty:
        return
    for (dt, key2), val in ser.items():
        agg[(dt, key2)][colname] += float(val)


# =========================== Aturan & Streaming ===========================
def _apply_rules_and_update_main(df_chunk: pd.DataFrame, agg) -> None:
    H = df_chunk[COL_H].fillna("").astype(str).str.lower()
    AA = df_chunk[COL_AA].fillna("").astype(str).str.lower()
    X = df_chunk[COL_X].fillna("").astype(str).str.lower()
    ASAL = df_chunk[COL_ASAL].fillna("Tidak diketahui").astype(str).str.strip()
    amt = pd.to_numeric(df_chunk[COL_K], errors="coerce").fillna(0)
    tgl = df_chunk["Tanggal"]

    def sum_by_key(mask) -> pd.Series:
        if mask.any():
            return amt[mask].groupby([tgl[mask], ASAL[mask]], dropna=False).sum(min_count=1)
        mi = pd.MultiIndex.from_arrays([[], []], names=["Tanggal", "Pelabuhan"])
        return pd.Series(index=mi, dtype="float64")

    rules = OrderedDict([
        ("Cash", H.str.contains("cash", na=False)),
        ("Prepaid BRI", H.str.contains("prepaid-bri", na=False)),
        ("Prepaid BNI", H.str.contains("prepaid-bni", na=False)),
        ("Prepaid Mandiri", H.str.contains("prepaid-mandiri", na=False)),
        ("Prepaid BCA", H.str.contains("prepaid-bca", na=False)),
        ("SKPT", H.str.contains("skpt", na=False)),
        ("IFCS", H.str.contains("ifcs", na=False)),
        ("Reedem", H.str.contains("reedem", na=False) | H.str.contains("redeem", na=False)),
        ("ESPAY", H.str.contains("finpay", na=False) & AA.str.startswith("esp", na=False)),
        ("Finnet", H.str.contains("finpay", na=False) & (~AA.str.startswith("esp", na=False))),
    ])
    for name, m in rules.items():
        _update_agg_series_key2(agg, sum_by_key(m), name)

    # BCA/NON BCA dari SOF ID
    is_finpay = H.str.contains("finpay", na=False)
    is_bca_tag = X.str.contains("vabcaespay", na=False) | X.str.contains("bluespay", na=False)
    _update_agg_series_key2(agg, sum_by_key(is_finpay & is_bca_tag), "BCA")
    _update_agg_series_key2(agg, sum_by_key(is_finpay & (~is_bca_tag)), "NON BCA")

def _flush_xlsx_batch_main(buf: List[List], year: int, month: int, agg) -> None:
    df = pd.DataFrame(buf, columns=[COL_H, COL_B, COL_AA, COL_K, COL_X, COL_ASAL])
    t = pd.to_datetime(df[COL_B], errors="coerce", dayfirst=True)
    mask = (t.dt.year == year) & (t.dt.month == month)
    if not mask.any(): return
    sub = df.loc[mask].copy()
    sub["Tanggal"] = t.loc[mask].dt.date
    _apply_rules_and_update_main(sub, agg)

def _process_csv_main_fast(data: bytes, year: int, month: int, agg) -> None:
    engine = "pyarrow" if _HAS_PYARROW else "c"
    itr = pd.read_csv(
        io.BytesIO(data),
        usecols=REQUIRED_COLS,
        chunksize=CSV_CHUNK_ROWS,
        dtype={COL_H: "string", COL_AA: "string", COL_X: "string", COL_ASAL: "string"},
        engine=engine,
    )
    for chunk in itr:
        t = pd.to_datetime(chunk[COL_B], errors="coerce", dayfirst=True)
        mask = (t.dt.year == year) & (t.dt.month == month)
        if not mask.any(): continue
        sub = chunk.loc[mask].copy()
        sub["Tanggal"] = t.loc[mask].dt.date
        _apply_rules_and_update_main(sub, agg)

def _process_xlsx_main_streaming(data: bytes, year: int, month: int, agg) -> None:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None: wb.close(); return
        name_to_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        if not all(c in name_to_idx for c in REQUIRED_COLS): wb.close(); return
        buf = []
        for r in rows:
            try:
                buf.append([
                    r[name_to_idx[COL_H]], r[name_to_idx[COL_B]], r[name_to_idx[COL_AA]],
                    r[name_to_idx[COL_K]], r[name_to_idx[COL_X]], r[name_to_idx[COL_ASAL]],
                ])
            except Exception:
                continue
            if len(buf) >= XLSX_BATCH_ROWS:
                _flush_xlsx_batch_main(buf, year, month, agg); buf.clear()
        if buf:
            _flush_xlsx_batch_main(buf, year, month, agg); buf.clear()
        wb.close()
    except Exception:
        try:
            df = pd.read_excel(io.BytesIO(data), sheet_name=0, usecols=REQUIRED_COLS)
        except Exception:
            return
        t = pd.to_datetime(df[COL_B], errors="coerce", dayfirst=True)
        mask = (t.dt.year == year) & (t.dt.month == month)
        if not mask.any(): return
        sub = df.loc[mask].copy()
        sub["Tanggal"] = t.loc[mask].dt.date
        _apply_rules_and_update_main(sub, agg)

def _build_result_from_agg_main(agg) -> pd.DataFrame:
    if not agg: return pd.DataFrame()
    rows: List[dict] = []
    for (dt, asal), bucket in agg.items():
        row = {"Tanggal": dt, "Pelabuhan": asal}
        for c in CAT_COLS: row[c] = bucket.get(c, 0.0)
        row["Total"] = sum(row[c] for c in CAT_COLS)
        bca = bucket.get("BCA", 0.0); nonbca = bucket.get("NON BCA", 0.0)
        row["BCA"] = bca; row["NON BCA"] = nonbca
        row["NON"] = sum(row[c] for c in NON_COMPONENTS)
        row["TOTAL"] = bca + nonbca + row["NON"]  # why: permintaan user
        row["Selisih"] = row["TOTAL"] - row["Total"]
        rows.append(row)
    df = pd.DataFrame(rows)
    if df.empty: return df
    df = df[["Tanggal", "Pelabuhan"] + CAT_COLS + ["Total", "BCA", "NON BCA", "NON", "TOTAL", "Selisih"]]
    return df.sort_values(["Pelabuhan", "Tanggal"]).reset_index(drop=True)


# =========================== Parallel & UI ===========================
def _process_single_main(b: bytes, filename: str, year: int, month: int) -> Dict:
    agg = _empty_agg_key2()
    low = filename.lower()
    if low.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(b)) as zf:
            for m in zf.infolist():
                if m.is_dir(): continue
                fname = m.filename.lower()
                if not fname.endswith(VALID_EXTS): continue
                content = zf.read(m)
                if fname.endswith((".xlsx", ".xls")): _process_xlsx_main_streaming(content, year, month, agg)
                else: _process_csv_main_fast(content, year, month, agg)
    elif low.endswith((".xlsx", ".xls")):
        _process_xlsx_main_streaming(b, year, month, agg)
    elif low.endswith(".csv"):
        _process_csv_main_fast(b, year, month, agg)
    return agg

def _parallel_process_main(files: List["st.runtime.uploaded_file_manager.UploadedFile"], year: int, month: int, max_workers: int) -> Dict:
    if not files: return {}
    workers = min(max_workers, max(1, len(files)))
    agg_total = _empty_agg_key2()
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = []
        for f in files:
            try: data = f.getvalue()
            except Exception: data = f.read()
            futs.append(ex.submit(_process_single_main, data, f.name, year, month))
        for fu in as_completed(futs):
            part = fu.result()
            agg_total = _merge_aggs(agg_total, part)
    return agg_total

def main():
    st.set_page_config(page_title="Rekonsiliasi Payment Report (Tabel Utama)", layout="wide")
    st.title("Rekonsiliasi Payment Report (Tabel Utama)")

    # Sidebar: filter & uploader
    today = date.today()
    years_options = list(range(today.year - 5, today.year + 6))
    year = st.sidebar.selectbox("Tahun", options=years_options, index=years_options.index(today.year))
    month_names = {1:"01 - Januari",2:"02 - Februari",3:"03 - Maret",4:"04 - April",5:"05 - Mei",6:"06 - Juni",7:"07 - Juli",8:"08 - Agustus",9:"09 - September",10:"10 - Oktober",11:"11 - November",12:"12 - Desember"}
    month = st.sidebar.selectbox("Bulan", options=list(range(1, 13)), index=today.month - 1, format_func=lambda m: month_names[m])
    highlight = st.sidebar.checkbox("Highlight kolom Selisih ≠ 0", value=True)
    max_workers = st.sidebar.slider("Paralel file (workers)", 1, 8, value=4)

    up_files_main = st.sidebar.file_uploader("Upload ZIP / Excel / CSV (Tabel Utama)", type=["zip", "xlsx", "xls", "csv"], accept_multiple_files=True)

    st.subheader(f"Hasil • Periode: {month_names[month]} {year}")
    if not up_files_main:
        st.info("Upload file di panel kiri.")
        return

    with st.spinner("Memproses (paralel + streaming)…"):
        agg_main = _parallel_process_main(up_files_main, year=year, month=month, max_workers=max_workers)
    result_main = _build_result_from_agg_main(agg_main)

    if result_main.empty:
        st.warning("Tidak ada data valid setelah filter.")
        return

    ports = list(result_main["Pelabuhan"].dropna().unique()); ports.sort()
    tabs = st.tabs(ports if ports else ["(Tidak ada Pelabuhan)"])
    for tab, port in zip(tabs, ports):
        with tab:
            st.markdown(f"**Pelabuhan: {port}**")
            df_show = result_main[result_main["Pelabuhan"] == port].copy()
            df_show["Tanggal"] = pd.to_datetime(df_show["Tanggal"]).dt.strftime("%d/%m/%Y")
            df_show = _add_subtotal_row(df_show)
            num_cols = df_show.select_dtypes(include="number").columns
            df_show[num_cols] = df_show[num_cols].round(0).astype("Int64")
            st.dataframe(_style_table(df_show, highlight=highlight), use_container_width=True)

    st.divider()
    st.subheader("Unduh (Gabungan)")
    export_df = result_main.copy()
    export_df["Tanggal"] = pd.to_datetime(export_df["Tanggal"]).dt.strftime("%d/%m/%Y")
    num_cols = export_df.select_dtypes(include="number").columns
    export_df[num_cols] = export_df[num_cols].round(0).astype("Int64")
    csv_bytes = export_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button("Unduh CSV", data=csv_bytes, file_name=f"rekonsiliasi_payment_{year}_{month:02d}_per_pelabuhan.csv", mime="text/csv")
    excel_bytes, engine_used, err_msg = _to_excel_bytes(export_df, sheet_name="Rekonsiliasi")
    if excel_bytes:
        st.download_button(f"Unduh Excel (.xlsx){' • ' + engine_used if engine_used else ''}", data=excel_bytes, file_name=f"rekonsiliasi_payment_{year}_{month:02d}_per_pelabuhan.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    elif err_msg:
        st.warning("Ekspor Excel dinonaktifkan. Tambahkan `xlsxwriter`/`openpyxl`.\n" + err_msg)


if __name__ == "__main__":
    main()


# ======================================================================
# FILE 2: streamlit_settlement_espay.py  (Settlement Dana ESPAY)
# ======================================================================
import io
import zipfile
from datetime import date
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from zoneinfo import ZoneInfo

# Optional akselerasi CSV
_HAS_PYARROW = False
try:
    import pyarrow  # noqa: F401
    _HAS_PYARROW = True
except Exception:
    _HAS_PYARROW = False


# =========================== Konfigurasi ===========================
ESPAY_DATE = "Settlement Date"   # kolom tanggal
ESPAY_PROD = "Product Name"      # VA vs non-VA
ESPAY_VA_NAME = "VA Name"        # mapping pelabuhan
ESPAY_REQUIRED = [ESPAY_DATE, ESPAY_PROD, ESPAY_VA_NAME]
ESPAY_AMOUNT_CANDIDATES = ["Amount", "Settlement Amount", "Amount (Rp.)", "Total Amount", "TOTAL", "Total"]

# Pola longgar VA Name → Pelabuhan (UPPERCASE)
PORT_PATTERNS = {
    "BAKAUHENI": ["asdp bakauheni", "bakauheni"],
    "MERAK": ["asdp merak", "merak"],
    "KETAPANG": ["asdp ketapang", "ketapang"],
    "GILIMANUK": ["asdp gilimanuk", "gilimanuk"],
}

CSV_CHUNK_ROWS = 300_000
XLSX_BATCH_ROWS = 80_000
VALID_EXTS = (".xlsx", ".xls", ".csv")


# =========================== Utilitas ===========================
def _style_table(df_display: pd.DataFrame) -> "pd.io.formats.style.Styler":
    numeric_cols = df_display.select_dtypes(include="number").columns.tolist()
    return df_display.style.format("{:,.0f}", subset=numeric_cols)

def _add_subtotal_row(df_display: pd.DataFrame, label: str = "Subtotal", date_col: str = "Tanggal") -> pd.DataFrame:
    # why: subtotal di UI
    numeric_cols = df_display.select_dtypes(include="number").columns.tolist()
    totals = df_display[numeric_cols].sum()
    subtotal = {c: (totals[c] if c in totals else None) for c in df_display.columns}
    subtotal[date_col] = label
    return pd.concat([df_display, pd.DataFrame([subtotal])], ignore_index=True)

def _to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Settlement ESPAY") -> Tuple[Optional[bytes], Optional[str], Optional[str]]:
    for engine in ("xlsxwriter", "openpyxl"):
        try:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine=engine) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            return buf.getvalue(), engine, None
        except ImportError:
            continue
        except Exception as e:
            return None, None, f"Gagal menulis Excel dengan {engine}: {e}"
    return None, None, "Tidak ada engine Excel (xlsxwriter/openpyxl)."

def _localize_to_tz(series_dt: pd.Series, tz_name: str) -> pd.Series:
    tz = ZoneInfo(tz_name)
    try:
        if series_dt.dt.tz is not None:
            return series_dt.dt.tz_convert(tz)
        return series_dt.dt.tz_localize(tz, nonexistent="shift_forward", ambiguous="NaT")
    except Exception:
        try:
            return series_dt.dt.tz_localize(tz, nonexistent="shift_forward", ambiguous="NaT")
        except Exception:
            return series_dt

def _detect_amount_col(cols: List[str]) -> Optional[str]:
    norm = {c.strip().lower(): c for c in cols}
    for cand in ESPAY_AMOUNT_CANDIDATES:
        if cand.lower() in norm: return norm[cand.lower()]
    return None

def _port_from_va_name(val: str) -> str:
    s = (val or "").strip().lower()
    for label, patterns in PORT_PATTERNS.items():
        for pat in patterns:
            if pat in s:  # why: toleran variasi penamaan VA Name
                return label
    return "LAINNYA"


# =========================== Agregator & Streaming ===========================
def _empty_agg():
    return defaultdict(lambda: defaultdict(float))  # key: (Tanggal, Pelabuhan) -> {Virtual Account/E-Money: sum}

def _update_agg(agg, ser: pd.Series, field: str) -> None:
    if ser.empty: return
    for (dt, port), val in ser.items():
        agg[(dt, port)][field] += float(val)

def _apply_and_update_espay(df_chunk: pd.DataFrame, agg, tz_name: str) -> None:
    cols_map = {c.strip(): c for c in df_chunk.columns}
    for need in ESPAY_REQUIRED:
        if need not in cols_map: return

    prod = df_chunk[cols_map[ESPAY_PROD]].fillna("").astype(str).str.lower()
    va_name = df_chunk[cols_map[ESPAY_VA_NAME]].fillna("").astype(str)

    t = pd.to_datetime(df_chunk[cols_map[ESPAY_DATE]], errors="coerce", dayfirst=True)
    t_local = _localize_to_tz(t, tz_name)
    df_chunk = df_chunk.copy()
    df_chunk["Tanggal"] = t_local.dt.date

    amt_col = _detect_amount_col(list(df_chunk.columns))
    amt = pd.to_numeric(df_chunk[amt_col], errors="coerce").fillna(0) if amt_col else pd.Series(1.0, index=df_chunk.index)
    port = va_name.map(_port_from_va_name)
    is_va = prod.str.contains("va", na=False)

    def sum_by(mask) -> pd.Series:
        if mask.any():
            return amt[mask].groupby([df_chunk["Tanggal"][mask], port[mask]], dropna=False).sum(min_count=1)
        mi = pd.MultiIndex.from_arrays([[], []], names=["Tanggal", "Pelabuhan"])
        return pd.Series(index=mi, dtype="float64")

    _update_agg(agg, sum_by(is_va), "Virtual Account")
    _update_agg(agg, sum_by(~is_va), "E-Money")

def _espay_header_usecols_csv(b: bytes) -> List[str]:
    engine = "pyarrow" if _HAS_PYARROW else "c"
    df0 = pd.read_csv(io.BytesIO(b), nrows=0, dtype="unicode", engine=engine)
    cols = [c.strip() for c in df0.columns]
    use = [c for c in cols if c in ESPAY_REQUIRED]
    amt = _detect_amount_col(cols)
    if amt: use.append(amt)
    return list(dict.fromkeys(use))

def _process_csv_espay_fast(b: bytes, year: int, month: int, agg, tz_name: str) -> Tuple[Optional[str], int, int]:
    usecols = _espay_header_usecols_csv(b)
    if not all(c in usecols for c in ESPAY_REQUIRED): return None, 0, 0
    engine = "pyarrow" if _HAS_PYARROW else "c"
    itr = pd.read_csv(io.BytesIO(b), chunksize=CSV_CHUNK_ROWS, dtype="unicode", usecols=usecols, engine=engine)
    amount_used = _detect_amount_col(usecols)
    total_rows = 0; used_rows = 0
    for chunk in itr:
        total_rows += len(chunk)
        chunk.columns = [str(c).strip() for c in chunk.columns]
        t = pd.to_datetime(chunk[ESPAY_DATE], errors="coerce", dayfirst=True)
        t_local = _localize_to_tz(t, tz_name)
        mask = (t_local.dt.year == year) & (t_local.dt.month == month)
        if not mask.any(): continue
        used_rows += int(mask.sum())
        sub = chunk.loc[mask].copy()
        sub[ESPAY_DATE] = t_local.loc[mask]
        _apply_and_update_espay(sub, agg, tz_name)
    return amount_used, total_rows, used_rows

def _flush_xlsx_batch_espay(buf: List[List], year: int, month: int, agg, amt_col_name: Optional[str], tz_name: str) -> int:
    cols = [ESPAY_PROD, ESPAY_DATE, ESPAY_VA_NAME, "__AMT__"]
    df = pd.DataFrame(buf, columns=cols)
    t = pd.to_datetime(df[ESPAY_DATE], errors="coerce", dayfirst=True)
    t_local = _localize_to_tz(t, tz_name)
    mask = (t_local.dt.year == year) & (t_local.dt.month == month)
    if not mask.any(): return 0
    sub = df.loc[mask, [ESPAY_PROD, ESPAY_DATE, ESPAY_VA_NAME, "__AMT__"]].copy()
    if "__AMT__" in sub.columns:
        sub.rename(columns={"__AMT__": ESPAY_AMOUNT_CANDIDATES[0]}, inplace=True)  # why: samarkan kolom amount agar terdeteksi
    _apply_and_update_espay(sub, agg, tz_name)
    return int(mask.sum())

def _process_xlsx_espay_streaming(b: bytes, year: int, month: int, agg, tz_name: str) -> Tuple[Optional[str], int, int]:
    rows_count = 0; rows_used = 0
    try:
        wb = load_workbook(io.BytesIO(b), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None: wb.close(); return None, 0, 0
        header = [str(h).strip() if h is not None else "" for h in header]
        name_to_idx = {h: i for i, h in enumerate(header) if h}
        if not all(c in name_to_idx for c in ESPAY_REQUIRED): wb.close(); return None, 0, 0
        amt_col_name = _detect_amount_col(header)
        buf = []; amount_used = amt_col_name
        for r in rows:
            rows_count += 1
            try:
                buf.append([
                    r[name_to_idx[ESPAY_PROD]],
                    r[name_to_idx[ESPAY_DATE]],
                    r[name_to_idx[ESPAY_VA_NAME]],
                    (r[name_to_idx[amt_col_name]] if amt_col_name and amt_col_name in name_to_idx else None),
                ])
            except Exception:
                continue
            if len(buf) >= XLSX_BATCH_ROWS:
                rows_used += _flush_xlsx_batch_espay(buf, year, month, agg, amt_col_name, tz_name); buf.clear()
        if buf:
            rows_used += _flush_xlsx_batch_espay(buf, year, month, agg, amt_col_name, tz_name); buf.clear()
        wb.close()
        return amount_used, rows_count, rows_used
    except Exception:
        try:
            df = pd.read_excel(io.BytesIO(b), sheet_name=0)
        except Exception:
            return None, 0, 0
        df.columns = [str(c).strip() for c in df.columns]
        if not all(c in df.columns for c in ESPAY_REQUIRED): return None, 0, 0
        t = pd.to_datetime(df[ESPAY_DATE], errors="coerce", dayfirst=True)
        t_local = _localize_to_tz(t, tz_name)
        mask = (t_local.dt.year == year) & (t_local.dt.month == month)
        used = int(mask.sum())
        if not used: return _detect_amount_col(list(df.columns)), len(df), 0
        sub = df.loc[mask, [ESPAY_PROD, ESPAY_DATE, ESPAY_VA_NAME] + ([c for c in df.columns if c in ESPAY_AMOUNT_CANDIDATES])].copy()
        sub[ESPAY_DATE] = t_local.loc[mask]
        _apply_and_update_espay(sub, agg, tz_name)
        return _detect_amount_col(list(df.columns)), len(df), used


# =========================== Parallel & UI ===========================
def _merge_aggs(a: Dict, b: Dict) -> Dict:
    for k, bucket in b.items():
        for col, val in bucket.items():
            a[k][col] += val
    return a

def _process_single_espay(b: bytes, filename: str, year: int, month: int, tz_name: str) -> Tuple[Dict, str]:
    agg = defaultdict(lambda: defaultdict(float))
    low = filename.lower()
    info = ""
    if low.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(b)) as zf:
            for m in zf.infolist():
                if m.is_dir(): continue
                fname = m.filename.lower()
                if not fname.endswith(VALID_EXTS): continue
                content = zf.read(m)
                if fname.endswith((".xlsx", ".xls")):
                    c, r_all, r_used = _process_xlsx_espay_streaming(content, year, month, agg, tz_name)
                else:
                    c, r_all, r_used = _process_csv_espay_fast(content, year, month, agg, tz_name)
                info += f"{filename}::{m.filename} → {'sum ' + str(c) if c else 'count trx'} (rows: {r_used}/{r_all})\n"
    elif low.endswith((".xlsx", ".xls")):
        c, r_all, r_used = _process_xlsx_espay_streaming(b, year, month, agg, tz_name)
        info += f"{filename} → {'sum ' + str(c) if c else 'count trx'} (rows: {r_used}/{r_all})\n"
    elif low.endswith(".csv"):
        c, r_all, r_used = _process_csv_espay_fast(b, year, month, agg, tz_name)
        info += f"{filename} → {'sum ' + str(c) if c else 'count trx'} (rows: {r_used}/{r_all})\n"
    return agg, info.strip()

def _parallel_process_espay(files: List["st.runtime.uploaded_file_manager.UploadedFile"], year: int, month: int, tz_name: str, max_workers: int) -> Tuple[Dict, List[str]]:
    if not files: return {}, []
    workers = min(max_workers, max(1, len(files)))
    agg_total = defaultdict(lambda: defaultdict(float))
    logs: List[str] = []
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = []
        for f in files:
            try: data = f.getvalue()
            except Exception: data = f.read()
            futs.append(ex.submit(_process_single_espay, data, f.name, year, month, tz_name))
        for fu in as_completed(futs):
            part_agg, info = fu.result()
            agg_total = _merge_aggs(agg_total, part_agg)
            if info: logs.append(info)
    return agg_total, logs

def _build_result_from_agg(agg) -> pd.DataFrame:
    if not agg: return pd.DataFrame()
    rows = []
    for (dt, port), bucket in agg.items():
        rows.append({
            "Tanggal": dt,
            "Pelabuhan": port,
            "Virtual Account": bucket.get("Virtual Account", 0.0),
            "E-Money": bucket.get("E-Money", 0.0),
        })
    df = pd.DataFrame(rows)
    if df.empty: return df
    df = df[["Tanggal", "Pelabuhan", "Virtual Account", "E-Money"]]
    return df.sort_values(["Pelabuhan", "Tanggal"]).reset_index(drop=True)

def main():
    st.set_page_config(page_title="Settlement Dana ESPAY", layout="wide")
    st.title("Settlement Dana ESPAY")

    # Sidebar: filter & uploader
    today = date.today()
    years_options = list(range(today.year - 5, today.year + 6))
    year = st.sidebar.selectbox("Tahun", options=years_options, index=years_options.index(today.year))
    month_names = {1:"01 - Januari",2:"02 - Februari",3:"03 - Maret",4:"04 - April",5:"05 - Mei",6:"06 - Juni",7:"07 - Juli",8:"08 - Agustus",9:"09 - September",10:"10 - Oktober",11:"11 - November",12:"12 - Desember"}
    month = st.sidebar.selectbox("Bulan", options=list(range(1, 13)), index=today.month - 1, format_func=lambda m: month_names[m])
    tz_choice = st.sidebar.selectbox("Zona waktu Settlement Date", options=["Asia/Jakarta", "UTC", "Asia/Makassar", "Asia/Jayapura"], index=0)
    max_workers = st.sidebar.slider("Paralel file (workers)", 1, 8, value=4)

    up_files = st.sidebar.file_uploader("Upload ZIP / Excel / CSV (ESPAY)", type=["zip", "xlsx", "xls", "csv"], accept_multiple_files=True)

    st.subheader(f"Hasil • Periode: {month_names[month]} {year}")
    if not up_files:
        st.info("Upload file di panel kiri.")
        return

    with st.spinner("Memproses Settlement ESPAY (paralel + streaming)…"):
        agg, logs = _parallel_process_espay(up_files, year=year, month=month, tz_name=tz_choice, max_workers=max_workers)

    result = _build_result_from_agg(agg)
    if result.empty:
        st.warning("Tidak ada data valid setelah filter.")
        if logs:
            with st.expander("Log pemrosesan"):
                st.code("\n".join(logs))
        return

    # Tabs per Pelabuhan (sama seperti tabel utama)
    ports = list(result["Pelabuhan"].dropna().unique()); ports.sort()
    tabs = st.tabs(ports if ports else ["(Tidak ada Pelabuhan)"])
    for tab, port in zip(tabs, ports):
        with tab:
            st.markdown(f"**Pelabuhan: {port}**")
            df_show = result[result["Pelabuhan"] == port].copy()
            df_show["Tanggal"] = pd.to_datetime(df_show["Tanggal"]).dt.strftime("%d/%m/%Y")
            df_show = _add_subtotal_row(df_show)
            num_cols = df_show.select_dtypes(include="number").columns
            df_show[num_cols] = df_show[num_cols].round(0).astype("Int64")
            st.dataframe(_style_table(df_show), use_container_width=True)

    st.divider()
    st.subheader("Unduh (Gabungan)")
    export_df = result.copy()
    export_df["Tanggal"] = pd.to_datetime(export_df["Tanggal"]).dt.strftime("%d/%m/%Y")
    num_cols = export_df.select_dtypes(include="number").columns
    export_df[num_cols] = export_df[num_cols].round(0).astype("Int64")
    csv_bytes = export_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button("Unduh CSV", data=csv_bytes, file_name=f"settlement_espay_{year}_{month:02d}.csv", mime="text/csv")
    excel_bytes, engine_used, err_msg = _to_excel_bytes(export_df)
    if excel_bytes:
        st.download_button(f"Unduh Excel (.xlsx){' • ' + engine_used if engine_used else ''}", data=excel_bytes, file_name=f"settlement_espay_{year}_{month:02d}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    elif err_msg:
        st.warning("Ekspor Excel dinonaktifkan. Tambahkan `xlsxwriter`/`openpyxl`.\n" + err_msg)

    if logs:
        with st.expander("Log pemrosesan"):
            st.code("\n".join(logs))


if __name__ == "__main__":
    main()
